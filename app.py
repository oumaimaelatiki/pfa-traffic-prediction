import streamlit as st
import pandas as pd
import plotly.express as px
from openpyxl import load_workbook
from groq import Groq

st.set_page_config(page_title="Trafic Casablanca", page_icon="car", layout="wide")

@st.cache_data
def load_data():
    wb = load_workbook('Dataset for traffic analysis in Casablanca, Morocco.xlsx', read_only=True)
    sheets = ['Table 5. Monday','Table 6. Tuesday','Table 7. Wednesday','Table. 8 Thursday','Table. 9 Friday','Table. 10 Saturday','Table. 11 Sunday']
    jours = ['Lundi','Mardi','Mercredi','Jeudi','Vendredi','Samedi','Dimanche']
    rows = []
    for sheet, jour in zip(sheets, jours):
        ws = wb[sheet]
        data = list(ws.iter_rows(values_only=True))
        header = None
        for i, row in enumerate(data):
            if row and any('COMMUNE' in str(c) for c in row if c):
                header = i
                break
        if header is None:
            continue
        for row in data[header+1:]:
            if row and row[1] and str(row[1]).strip() not in ['', 'None'] and str(row[1]).strip() != 'COMMUNE':
                commune = str(row[1]).strip()
                for h in range(24):
                    try:
                        val = row[12+h]
                        if val and str(val) != 'None' and float(val) < 100:
                            rows.append({'commune': commune, 'jour': jour, 'heure': h, 'temps_trajet': float(val)})
                    except:
                        pass
    return pd.DataFrame(rows)

df = load_data()
communes = sorted(df['commune'].unique()) if len(df) > 0 else []
jours = ['Lundi','Mardi','Mercredi','Jeudi','Vendredi','Samedi','Dimanche']

page = st.sidebar.selectbox("Navigation", ["Accueil", "Analyse du trafic", "Prediction", "Assistant IA"])

if page == "Accueil":
    st.title("Application de Trafic Routier - Casablanca")
    st.markdown("### Bienvenue dans l application d analyse et prediction du trafic de Casablanca")
    col1, col2, col3 = st.columns(3)
    col1.metric("Communes", "22")
    col2.metric("Jours analyses", "7")
    col3.metric("Heures/jour", "24")
    st.markdown("---")
    st.info("Cette application analyse le trafic routier de Casablanca en utilisant des donnees collectees via l API Waze.")
    st.markdown("#### Pages disponibles")
    st.markdown("- **Analyse du trafic** : Visualisation des donnees de trafic par commune et heure")
    st.markdown("- **Prediction** : Prediction du temps de trajet selon la commune et l heure")
    st.markdown("- **Assistant IA** : Posez n importe quelle question sur le trafic de Casablanca")

elif page == "Analyse du trafic":
    st.title("Analyse du Trafic - Casablanca")
    if len(df) == 0:
        st.error("Donnees non chargees.")
    else:
        col1, col2 = st.columns(2)
        commune_sel = col1.selectbox("Commune", ["Toutes"] + list(communes))
        jour_sel = col2.selectbox("Jour", ["Tous"] + jours)
        dff = df.copy()
        if commune_sel != "Toutes":
            dff = dff[dff['commune'] == commune_sel]
        if jour_sel != "Tous":
            dff = dff[dff['jour'] == jour_sel]
        st.markdown("---")
        fig1 = px.line(dff.groupby('heure')['temps_trajet'].mean().reset_index(),
            x='heure', y='temps_trajet',
            title='Temps de trajet moyen par heure',
            labels={'heure':'Heure','temps_trajet':'Temps (min)'},
            color_discrete_sequence=['#e74c3c'])
        st.plotly_chart(fig1, use_container_width=True)
        fig2 = px.bar(dff.groupby('commune')['temps_trajet'].mean().reset_index().sort_values('temps_trajet', ascending=False),
            x='commune', y='temps_trajet',
            title='Temps de trajet moyen par commune',
            labels={'commune':'Commune','temps_trajet':'Temps (min)'},
            color_discrete_sequence=['#3498db'])
        fig2.update_xaxes(tickangle=45)
        st.plotly_chart(fig2, use_container_width=True)
        pivot = dff.pivot_table(values='temps_trajet', index='jour', columns='heure', aggfunc='mean')
        fig3 = px.imshow(pivot, title='Carte thermique - Trafic par jour et heure',
            color_continuous_scale='YlOrRd',
            labels={'x':'Heure','y':'Jour','color':'Temps (min)'})
        st.plotly_chart(fig3, use_container_width=True)

elif page == "Prediction":
    st.title("Prediction du Temps de Trajet")
    if len(df) == 0:
        st.error("Donnees non chargees.")
    else:
        col1, col2, col3 = st.columns(3)
        commune_sel = col1.selectbox("Commune", communes)
        jour_sel = col2.selectbox("Jour", jours)
        heure_sel = col3.slider("Heure", 0, 23, 8)
        if st.button("Predire", type="primary"):
            filtre = df[(df['commune']==commune_sel) & (df['jour']==jour_sel) & (df['heure']==heure_sel)]
            if len(filtre) > 0:
                pred = filtre['temps_trajet'].mean()
                st.success(f"Temps de trajet predit : {pred:.1f} minutes")
                moy_generale = df[df['heure']==heure_sel]['temps_trajet'].mean()
                diff = ((pred - moy_generale) / moy_generale) * 100
                if diff > 10:
                    st.warning(f"Trafic plus charge que la moyenne ({diff:.1f}%)")
                elif diff < -10:
                    st.info(f"Trafic plus fluide que la moyenne ({abs(diff):.1f}% moins charge)")
                else:
                    st.info("Trafic normal pour cette heure")
                fig = px.line(df[df['commune']==commune_sel].groupby('heure')['temps_trajet'].mean().reset_index(),
                    x='heure', y='temps_trajet',
                    title=f'Profil du trafic - {commune_sel}',
                    labels={'heure':'Heure','temps_trajet':'Temps (min)'})
                fig.add_vline(x=heure_sel, line_dash="dash", line_color="red",
                    annotation_text=f"{heure_sel}h selectionnee")
                st.plotly_chart(fig, use_container_width=True)
            else:
                st.error("Pas de donnees pour cette selection.")

elif page == "Assistant IA":
    st.title("Assistant IA - Trafic Casablanca")
    st.markdown("Posez **n importe quelle question** sur le trafic de Casablanca !")
    GROQ_API_KEY = "VOTRE_CLE_API_ICI"        commune_max = df.groupby("commune")["temps_trajet"].mean().idxmax()
        commune_min = df.groupby("commune")["temps_trajet"].mean().idxmin()
        heure_max = int(df.groupby("heure")["temps_trajet"].mean().idxmax())
        heure_min = int(df.groupby("heure")["temps_trajet"].mean().idxmin())
        moy = df["temps_trajet"].mean()
        trafic_we = df[df["jour"].isin(["Samedi","Dimanche"])]["temps_trajet"].mean()
        trafic_sem = df[~df["jour"].isin(["Samedi","Dimanche"])]["temps_trajet"].mean()
        communes_list = ", ".join(sorted(df["commune"].unique()))
        stats_communes = df.groupby("commune")["temps_trajet"].mean().round(1).to_dict()
        stats_str = ", ".join([f"{k}: {v} min" for k, v in stats_communes.items()])
        context = f"""Tu es un assistant expert en trafic routier de Casablanca, Maroc.
Voici les donnees reelles du trafic de Casablanca collectees via Waze:
- Communes analysees: {communes_list}
- Temps moyen par commune: {stats_str}
- Commune avec le plus de trafic: {commune_max}
- Commune la plus fluide: {commune_min}
- Heure de pointe: {heure_max}h
- Meilleure heure pour voyager: {heure_min}h
- Temps de trajet moyen general: {moy:.1f} minutes
- Trafic semaine: {trafic_sem:.1f} min | Weekend: {trafic_we:.1f} min
Reponds toujours en francais. Sois precis et utilise les donnees ci-dessus pour repondre."""
    else:
        context = "Tu es un assistant expert en trafic routier de Casablanca. Reponds en francais."
    if "messages" not in st.session_state:
        st.session_state.messages = []
    for msg in st.session_state.messages:
        with st.chat_message(msg["role"]):
            st.write(msg["content"])
    question = st.chat_input("Posez n importe quelle question sur le trafic de Casablanca...")
    if question:
        st.session_state.messages.append({"role": "user", "content": question})
        with st.chat_message("user"):
            st.write(question)
        with st.spinner("Agent IA reflechit..."):
            try:
                messages_groq = [{"role": "system", "content": context}]
                for msg in st.session_state.messages[-6:]:
                    messages_groq.append({"role": msg["role"], "content": msg["content"]})
                response = client.chat.completions.create(
                    model="llama-3.3-70b-versatile",
                    messages=messages_groq,
                    max_tokens=500,
                    temperature=0.7
                )
                reponse = response.choices[0].message.content
            except Exception as e:
                reponse = f"Erreur: {str(e)}"
        st.session_state.messages.append({"role": "assistant", "content": reponse})
        with st.chat_message("assistant"):
            st.write(reponse)
